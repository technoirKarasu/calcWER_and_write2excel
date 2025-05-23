import glob
import jiwer
from janome.tokenizer import Tokenizer
import openpyxl
import re


CORRECT_TXT_PATH = r"original.txt"       # This txt file is referenced as correct.
TARGET_EXCEL_PATH = r"STT_research.xlsx"      # This excel file is written WER into. 
TARGET_WORKSHEET_NAME = 'Service_A'     # 'Service_A' or 'Service_B'

def main():
    org_file_path = CORRECT_TXT_PATH
    original_file = preprocess_and_wakati(org_file_path)
    
    # Specify an excel file
    target_excel = TARGET_EXCEL_PATH
    try:
        workbook = openpyxl.load_workbook(target_excel)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        print("指定されたExcelファイルが見つかりませんでした。新しいファイルを作成しました。")
        
    # Specify the work sheet to write into
    target_worksheet = workbook[TARGET_WORKSHEET_NAME]

    for stt_file_path in glob.glob(r'*_output.txt'):
        stt_file = preprocess_and_wakati(stt_file_path)

        # Calculate the WER results
        output = jiwer.process_words(original_file, stt_file)
        visualize_result = jiwer.visualize_alignment(output)
        
        result_to_excel(target_worksheet, stt_file_path, visualize_result)

        # Save excel every writing (Unless you do this, data aren't saved in the excel)
        workbook.save(target_excel)


# Japanese text preprocessing and word segmentation
def wakati(text):
    tokenizer = Tokenizer()
    tokens = tokenizer.tokenize(text)
    words = [token.surface for token in tokens]
    return ' '.join(words)

# Execute wakati line by line
def preprocess_and_wakati(target_file):
    with open(target_file, mode='r', encoding="utf-8") as f:
        target = [wakati(line.rstrip('\n')) for line in f]  # Read including blank lines
    return target

# Get a specified number in the jiwer result
def extract_result_number(keyword, visualize_result):
    if keyword=="wer":
        match = re.search(r'wer=(\d+\.\d+)%', visualize_result)
    elif keyword=="number of sentences":
        match = re.search(r'number of sentences: (\d+)', visualize_result)
    else:
        match = re.search(rf'{keyword}=(\d+)', visualize_result)

    if match:
        match_value = match.group(1)  # Get the matching number
        return match_value 
    else:
        print(f"Not found the value of {keyword}") 
        
# Get the column number based on the specified word
def get_column_number(search_word, work_sheet):
    for i in range(1, work_sheet.max_column + 1):
        header_cell = work_sheet.cell(row=1, column=i).value
        if header_cell == search_word:
            return i
    return print(f"Not found the column number of {search_word}")

# Write the specified value into the worksheet
def write_specified_value(specified_value, target_worksheet, target_row_number, result_block):
    # Get the column number to write into
    target_column_number = get_column_number(specified_value, target_worksheet)
    
    if specified_value == 'Result-jiwer (WER)':
         extracted_value = result_block[-121:]
    else:
        extracted_value = extract_result_number(specified_value, result_block)
    
    # Show a detail of processing
    # print(f"[DEBUG] Write {specified_value=} -> {extracted_value=} into (row {target_row_number}, col {target_column_number})")
    
    # Write the value into the cell
    target_worksheet.cell(row=target_row_number, column=target_column_number).value = extracted_value

def result_to_excel(work_sheet, stt_file_path, visualize_result):
    # Get the row number from the head of the target file name
    target_row_number = int(stt_file_path[:2]) + 1
    
    # List values to be written
    list_of_values_written = ['wer', 'substitutions', 'deletions', 'insertions', 'hits', 'Result-jiwer (WER)']
    
    # Write values into excel
    for specified_value in list_of_values_written:
        write_specified_value(specified_value, work_sheet, target_row_number, visualize_result)

    print(f"Complete to write the row No.{target_row_number}")
    return

if __name__ == '__main__':
    main()
    


'''
参考URL

jiwerの関数についての詳細： https://jitsi.github.io/jiwer/reference/process/#process.CharacterOutput
ファイルの書き込み時にutf-8を指定する理由 https://stackoverflow.com/questions/27092833/unicodeencodeerror-charmap-codec-cant-encode-characters
ファイル名の文字列からOutputを削除 https://note.nkmk.me/python-str-remove-strip/
'''
